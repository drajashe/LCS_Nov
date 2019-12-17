# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *
# globally declare wb and sheet variable

# opening the existing excel file
wb = load_workbook('data/DETAILS_EMP.xlsx')

# create the sheet object
sheet = wb.active


def excel():
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 50
    sheet.column_dimensions['E'].width = 50
    sheet.column_dimensions['F'].width = 50
    sheet.column_dimensions['G'].width = 50
    sheet.column_dimensions['H'].width = 50
    sheet.column_dimensions['I'].width = 50
    sheet.column_dimensions['J'].width = 50
    sheet.column_dimensions['K'].width = 50

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Designation"
    sheet.cell(row=1, column=3).value = "Role"
    sheet.cell(row=1, column=4).value = "Reliability"
    sheet.cell(row=1, column=5).value = "Responsibility"
    sheet.cell(row=1, column=6).value = "Accountability"
    sheet.cell(row=1, column=7).value = "Creativity"
    sheet.cell(row=1, column=8).value = "Quality"
    sheet.cell(row=1, column=9).value = "Speed"
    sheet.cell(row=1, column=10).value = "Expression"



# Function to set focus (cursor)
def focus1(event):
    # set focus on the course_field box
    designation_field.focus_set()


# Function to set focus
def focus2(event):
    # set focus on the sem_field box
    role_field.focus_set()


# Function to set focus
def focus3(event):
    # set focus on the form_no_field box
    reliability_field.focus_set()


# Function to set focus
def focus4(event):
    # set focus on the contact_no_field box
    responsibility_field.focus_set()


# Function to set focus
def focus5(event):
    # set focus on the email_id_field box
    accountability_field.focus_set()


# Function to set focus
def focus6(event):
    # set focus on the address_field box
    creativity_field.focus_set()

# Function to set focus
def focus7(event):
    # set focus on the address_field box
    quality_field.focus_set()

# Function to set focus
def focus8(event):
    # set focus on the address_field box
    speed_field.focus_set()

# Function to set focus
def focus9(event):
    # set focus on the address_field box
    expression_field.focus_set()

# Function to set focus
def focus10(event):
    # set focus on the address_field box
    posturing_field.focus_set()

# # Function to set focus
# def focus11(event):
#     # set focus on the address_field box
#     all_field.focus_set()




# Function for clearing the
# contents of text entry boxes
def clear():
    # clear the content of text entry box
    name_field.delete(0, END)
    designation_field.delete(0, END)
    role_field.delete(0, END)
    reliability_field.delete(0, END)
    responsibility_field.delete(0, END)
    accountability_field.delete(0, END)
    creativity_field.delete(0, END)
    quality_field.delete(0, END)
    speed_field.delete(0, END)
    expression_field.delete(0, END)
    posturing_field.delete(0, END)
    # overall_field.delete(0, END)





# Function to take data from GUI
# window and write to an excel file
def insert():
    # if user not fill any entry
    # then print "empty input"
    if (name_field.get() == "" and
            designation_field.get() == "" and
            role_field.get() == "" and
            reliability_field.get() == "" and
            responsibility_field.get() == "" and
            accountability_field.get() == "" and
            creativity_field.get() == "" and
            quality_field.get()=="" and
            speed_field.get()=="" and
            expression_field.get()=="" and
            posturing_field.get()==""
            # overall_field.get()==""
        ):
        print("empty input")

    else:

        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the ,
        current_row = sheet.max_row
        current_column = sheet.max_column

        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = designation_field.get()
        sheet.cell(row=current_row + 1, column=3).value = role_field.get()
        sheet.cell(row=current_row + 1, column=4).value = reliability_field.get()
        sheet.cell(row=current_row + 1, column=5).value = responsibility_field.get()
        sheet.cell(row=current_row + 1, column=6).value = accountability_field.get()
        sheet.cell(row=current_row + 1, column=7).value = creativity_field.get()
        sheet.cell(row=current_row + 1, column=8).value = quality_field.get()
        sheet.cell(row=current_row + 1, column=9).value = speed_field.get()
        sheet.cell(row=current_row + 1, column=10).value = expression_field.get()
        sheet.cell(row=current_row + 1, column=11).value = posturing_field.get()
        sheet.cell(row=current_row + 1, column=12).value = overall_field.get()


        # save the file
        wb.save('data/DETAILS_EMP.xlsx')

        # set focus on the name_field box
        name_field.focus_set()

        # call the clear() function
        clear()

    # Driver code


if __name__ == "__main__":
    # create a GUI window
    root = Tk()

    # set the background colour of GUI window
    root.configure(background='light blue')

    # set the title of GUI window
    root.title("Employee Feedback form")

    # set the configuration of GUI window
    root.geometry("700x600")

    excel()


    # create a Form label
    heading = Label(root, text="Form", bg="light grey")

    # create a Name label
    name = Label(root, text="Name", bg="light grey")

    # create a  designation label
    designation = Label(root, text="Designation", bg="light grey")

    # create a role label
    role = Label(root, text="Role", bg="light grey")

    # create a reliability lablel
    reliability = Label(root, text="Reliability", bg="light grey")

    # create a Responsibility label
    responsibility = Label(root, text="Responsibility", bg="light grey")

    # create a accountability label
    accountability = Label(root, text="Accountability", bg="light grey")

    # create a creativity label
    creativity = Label(root, text="Creativity", bg="light grey")

    # create a quality label
    quality = Label(root, text="Quality", bg="light grey")

    # create a speed label
    speed = Label(root, text="Speed", bg="light grey")

    # create a Expression label
    expression = Label(root, text="Expression", bg="light grey")

    # create a posturing label
    posturing = Label(root, text="Posturing", bg="light grey")

    # create a Overall label
    overall = Label(root, text="Overall", bg="light grey")

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    designation.grid(row=2, column=0)
    role.grid(row=3, column=0)
    reliability.grid(row=4, column=0)
    responsibility.grid(row=5, column=0)
    accountability.grid(row=6, column=0)
    creativity.grid(row=7, column=0)
    quality.grid(row=8, column=0)
    speed.grid(row=9, column=0)
    expression.grid(row=10, column=0)
    posturing.grid(row=11, column=0)
    overall.grid(row=12, column=0)





    # create a text entry box
    # for typing the information
    name_field = Entry(root)
    designation_field = Entry(root)
    role_field = Entry(root)
    reliability_field = Entry(root)
    responsibility_field = Entry(root)
    accountability_field = Entry(root)
    creativity_field = Entry(root)
    quality_field = Entry(root)
    speed_field = Entry(root)
    expression_field = Entry(root)
    posturing_field = Entry(root)
    overall_field = Entry(root)




    # bind method of widget is used for
    # the binding the function with the events

    # whenever the enter key is pressed
    # then call the focus1 function
    name_field.bind("<Return>", focus1)

    # whenever the enter key is pressed
    # then call the focus2 function
    designation_field.bind("<Return>", focus2)

    # whenever the enter key is pressed
    # then call the focus3 function
    role_field.bind("<Return>", focus3)

    # whenever the enter key is pressed
    # then call the focus4 function
    reliability_field.bind("<Return>", focus4)

    # whenever the enter key is pressed
    # then call the focus5 function
    responsibility_field.bind("<Return>", focus5)

    # whenever the enter key is pressed
    # then call the focus6 function
    accountability_field.bind("<Return>", focus6)

    # whenever the enter key is pressed
    # then call the focus6 function
    creativity_field.bind("<Return>", focus7)

    # whenever the enter key is pressed
    # then call the focus6 function
    quality_field.bind("<Return>", focus8)

    # whenever the enter key is pressed
    # then call the focus6 function
    speed_field.bind("<Return>", focus9)

    # whenever the enter key is pressed
    # then call the focus6 function
    expression_field.bind("<Return>", focus10)

    # whenever the enter key is pressed
    # then call the focus6 function
   # posturing_field.bind("<Return>", focus11)


    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    name_field.grid(row=1, column=1, ipadx="100")
    designation_field.grid(row=2, column=1, ipadx="100")
    role_field.grid(row=3, column=1, ipadx="100")
    reliability_field.grid(row=4, column=1, ipadx="100")
    responsibility_field.grid(row=5, column=1, ipadx="100")
    accountability_field.grid(row=6, column=1, ipadx="100")
    creativity_field.grid(row=7, column=1, ipadx="100")
    quality_field.grid(row=8, column=1, ipadx="100")
    speed_field.grid(row=9, column=1, ipadx="100")
    expression_field.grid(row=10, column=1, ipadx="100")
    posturing_field.grid(row=11, column=1, ipadx="100")
    # overall_field.grid(row=12, column=1, ipadx="100")


    # call excel function
    excel()

    # create a Submit Button and place into the root window
    submit = Button(root, text="Submit", fg="Black",
                    bg="Red", command=insert)
    submit.grid(row=15, column=1)

    # start the GUI
    root.mainloop()